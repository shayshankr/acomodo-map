#!/usr/bin/env python3
"""Generate data/portfolio-london.csv from the London tab of the XLSX export.

    python scripts/london_to_csv.py

The London tab has a different column layout from Ireland (a "Rooms size"
column, only three bed columns, GBP rents stored as bare numbers, a date-typed
move-in). This maps it — by header name, so small sheet edits don't break it —
onto the same canonical columns build_data.py already reads, injects the £
symbol (Google stores the currency as formatting, which the XLSX drops), and
tidies the float/date artifacts openpyxl returns.
"""

import csv
import re
import sys
from datetime import datetime, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Needs openpyxl:  pip install -r requirements.txt")

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "portfolio.xlsx"
OUT = ROOT / "data" / "portfolio-london.csv"
LONDON_TAB = "London Sept26 Portfolio details"

# Canonical header build_data.py expects (same as the Ireland tab).
CANONICAL = [
    "Property Name", "City", "Floor", "Room #", "Room Type", "Bedspaces",
    "# of washroom", "Monthly Rent\n(Per Bedspace)", "Monthly rent (For entire Unit)",
    "Tenant \nDemography", "Bed 1", "Bed 2", "Bed 3", "Bed 4", "Furnished",
    "Utilities", "Move in Date", "Payment terms", "Available Tenancies",
    "Media Link", "Key Features",
]

# Which London header (substring, lowercased) feeds each canonical column.
# None = leave blank (London has no equivalent).
SOURCE = {
    "Property Name": "property address",
    "City": "city",
    "Floor": "floor",
    "Room #": "room #",
    "Room Type": "room type",
    "Bedspaces": "bedspaces",
    "# of washroom": "# of washroom",
    "Monthly Rent\n(Per Bedspace)": "monthly rent",
    "Monthly rent (For entire Unit)": None,
    "Tenant \nDemography": "tenant",
    "Bed 1": "bed 1",
    "Bed 2": "bed 2",
    "Bed 3": "bed 3",
    "Bed 4": None,
    "Furnished": "furnished",
    "Utilities": "utility",
    "Move in Date": "move in",
    "Payment terms": "payment terms",
    "Available Tenancies": "available tenancies",
    "Media Link": "media link",
    "Key Features": "key features",
}

MONEY_COL = "Monthly Rent\n(Per Bedspace)"


def norm(text):
    return re.sub(r"\s+", " ", str(text or "")).strip().lower()


def find_columns(header):
    """Map each London header substring to its column index (first match wins)."""
    index = {}
    for i, cell in enumerate(header):
        index.setdefault(norm(cell), i)
    resolved = {}
    for canon, needle in SOURCE.items():
        if needle is None:
            resolved[canon] = None
            continue
        hit = next((i for key, i in index.items() if needle in key), None)
        resolved[canon] = hit
    return resolved


def tidy(value, *, is_money=False):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%-d %b %Y") if sys.platform != "win32" else value.strftime("%#d %b %Y")
    text = str(value)
    # "1.0" / "550.0" -> "1" / "550"
    if re.fullmatch(r"-?\d+\.0", text):
        text = text[:-2]
    if is_money and re.fullmatch(r"\d+(\.\d+)?", text):
        text = f"£{int(float(text)):,}"
    return text


def main():
    if not XLSX.exists():
        sys.exit(f"Missing {XLSX}. Download the sheet as .xlsx first.")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    if LONDON_TAB not in wb.sheetnames:
        sys.exit(f"No '{LONDON_TAB}' tab in the workbook.")
    ws = wb[LONDON_TAB]

    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    cols = find_columns(header)
    missing = [c for c, i in cols.items() if i is None and SOURCE[c] is not None]
    if missing:
        print(f"  note: no London column matched for {missing} (left blank)")

    out_rows = []
    for raw in rows[1:]:
        if not any(str(c).strip() for c in raw if c is not None):
            out_rows.append([""] * len(CANONICAL))
            continue
        line = []
        for canon in CANONICAL:
            idx = cols[canon]
            val = raw[idx] if idx is not None and idx < len(raw) else ""
            line.append(tidy(val, is_money=(canon == MONEY_COL)))
        out_rows.append(line)

    with OUT.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(CANONICAL)
        writer.writerows(out_rows)

    props = {r[0] for r in out_rows if r[0].strip()}
    print(f"Wrote {OUT.relative_to(ROOT)}: {len(props)} London properties, {sum(1 for r in out_rows if any(r))} rows")
    for p in sorted(props):
        print(f"  {p[:55]}")


if __name__ == "__main__":
    main()
